"""Parse PETROL FORMAT JUNE DELIVARY BOYS.xlsx → JSON for import.

Takes only: date, employee (via sheet/name map), kms, petrol amount.
Multiple slips on the same day for one employee are summed.
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip3 install --user openpyxl", file=sys.stderr)
    sys.exit(1)

# Sheet tab → employee email (Nithin tabs are disambiguated by sheet name).
SHEET_TO_EMAIL = {
    "SURYA": "surya.jillala@malconnexus.com",
    "VINITH": "vinithgoud.burlawar@malconnexus.com",
    "SHIVAJI": "shivaji.bashaboina@malconnexus.com",
    "SAI KRISHNA": "saikrishna.jeripothula@malconnexus.com",
    "NITHIN": "nithin.jatoth@malconnexus.com",
    "T NITHIN": "nithin.thaduri@malconnexus.com",
    "SWAMY": "swamy.katla@malconnexus.com",
    "G.PRAVEEN": "praveen.gandamalla@malconnexus.com",
    "SIDHAR": "siddhartha.kallepelly@malconnexus.com",
}

# Used on multi-name sheets like Driving Dept.
NAME_TO_EMAIL = {
    "SURYA": "surya.jillala@malconnexus.com",
    "VINEETH": "vinithgoud.burlawar@malconnexus.com",
    "VINITH": "vinithgoud.burlawar@malconnexus.com",
    "SHIVAJI": "shivaji.bashaboina@malconnexus.com",
    "SAI KRISHNA": "saikrishna.jeripothula@malconnexus.com",
    "SWAMY": "swamy.katla@malconnexus.com",
    "G.PRAVEEN": "praveen.gandamalla@malconnexus.com",
    "G PRAVEEN": "praveen.gandamalla@malconnexus.com",
    "SIDHARTH": "siddhartha.kallepelly@malconnexus.com",
    "SIDDARTHA": "siddhartha.kallepelly@malconnexus.com",
    "G.ARUN": "arunkumar.gopu@malconnexus.com",
    "G ARUN": "arunkumar.gopu@malconnexus.com",
    "D.SRINU": "srinu.thamadi@malconnexus.com",
    "D SRINU": "srinu.thamadi@malconnexus.com",
}

SKIP_SHEETS = {"SHEET1", "SHEET2", "VAMSHI"}  # empty / unused


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).upper()


def parse_kms(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"(\d+(?:\.\d+)?)", str(v))
    return float(m.group(1)) if m else 0.0


def parse_amount(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"(\d+(?:\.\d+)?)", str(v))
    return float(m.group(1)) if m else 0.0


def date_key(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return None


def resolve_email(sheet: str, emp_name: str) -> str | None:
    sheet_key = norm(sheet)
    if sheet_key in SHEET_TO_EMAIL:
        return SHEET_TO_EMAIL[sheet_key]
    name_key = norm(emp_name)
    if name_key in NAME_TO_EMAIL:
        return NAME_TO_EMAIL[name_key]
    return None


def main() -> None:
    path = Path(sys.argv[1] if len(sys.argv) > 1 else "PETROL FORMAT JUNE DELIVARY BOYS.xlsx")
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    # key: email|date → totals
    agg: dict[str, dict] = {}
    skipped: list[dict] = []
    raw_rows = 0

    for sheet in wb.sheetnames:
        if norm(sheet) in SKIP_SHEETS:
            continue
        ws = wb[sheet]
        for r in range(2, (ws.max_row or 0) + 1):
            dt = ws.cell(r, 3).value
            emp = ws.cell(r, 7).value
            kms = parse_kms(ws.cell(r, 9).value)
            amt = parse_amount(ws.cell(r, 10).value)
            dk = date_key(dt)
            if dk is None and amt == 0 and kms == 0:
                continue
            raw_rows += 1
            if dk is None:
                skipped.append({"sheet": sheet, "row": r, "reason": "missing date", "emp": emp})
                continue
            if amt == 0 and kms == 0:
                skipped.append({"sheet": sheet, "row": r, "reason": "no kms or amount", "emp": emp, "date": dk})
                continue
            email = resolve_email(sheet, emp or "")
            if not email:
                skipped.append({"sheet": sheet, "row": r, "reason": "unmapped employee", "emp": emp, "date": dk})
                continue
            key = f"{email}|{dk}"
            row = agg.get(key)
            if not row:
                row = {"email": email, "expenseDate": dk, "kmsDriven": 0.0, "petrolAmount": 0.0, "slips": 0}
                agg[key] = row
            row["kmsDriven"] += kms
            row["petrolAmount"] += amt
            row["slips"] += 1

    entries = sorted(agg.values(), key=lambda e: (e["expenseDate"], e["email"]))
    print(
        json.dumps(
            {
                "source": str(path.name),
                "rawRows": raw_rows,
                "entries": entries,
                "skipped": skipped,
                "unmappedEmployees": sorted(
                    {s.get("emp") for s in skipped if s.get("reason") == "unmapped employee" and s.get("emp")}
                ),
            },
            default=str,
        )
    )


if __name__ == "__main__":
    main()
